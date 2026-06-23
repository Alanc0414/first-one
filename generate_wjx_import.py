# -*- coding: utf-8 -*-
"""
生成问卷星官方 Excel 导入模板格式（7列，与平台下载模板一致）
列：题干 | 选项A | 选项B | 选项C | 选项D | 正确答案 | 解析
"""

import re
from openpyxl import Workbook

from generate_quiz import day1_data, day2_data, day3_data

WJX_COLUMNS = ['题干', '选项A', '选项B', '选项C', '选项D', '正确答案', '解析']


def clean_text(text):
    """去除问卷星不兼容的符号"""
    if not text:
        return ''
    text = str(text)
    text = text.replace('"', '').replace('"', '').replace('"', '')
    text = text.replace('（多选）', '').replace('(多选)', '')
    text = text.replace('[多选题]', '').replace('[单选题]', '')
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def parse_options(options_str):
    opts = []
    for line in options_str.strip().split('\n'):
        m = re.match(r'^[A-Z]\.\s*(.+)', line.strip())
        if m:
            opts.append(clean_text(m.group(1)))
    return opts


def convert_fill_question(question, answer_str):
    answers = [a.strip() for a in answer_str.replace('；', ';').split(';')]
    result = question
    for ans in answers:
        result = result.replace('______', '{' + ans + '}', 1)
    return clean_text(result), ';'.join(answers)


def convert_row(row):
    _, q_type, question, options_str, answer, source, explanation = row
    question = clean_text(question)
    analysis = clean_text(f'依据：{source}。{explanation}')

    if q_type == '填空题':
        question, fill_answer = convert_fill_question(question, answer)
        opts = ['', '', '', '']
        correct = fill_answer
    else:
        opts = parse_options(options_str)
        opts = (opts + ['', '', '', ''])[:4]
        correct = answer
        # 多选题：按问卷星规则，在题干末尾用括号标注正确答案，如 (ABC)
        if q_type == '多选题':
            question = re.sub(r'[？?]+$', '', question) + f'？({answer})'

    return [question] + opts + [correct, analysis]


def create_wjx_excel(filename, data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    for col_idx, col_name in enumerate(WJX_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(data, 2):
        converted = convert_row(row)
        for col_idx, val in enumerate(converted, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(filename)
    print('已生成：' + filename)


if __name__ == '__main__':
    create_wjx_excel('CMport_问卷星导入_Day1.xlsx', day1_data)
    create_wjx_excel('CMport_问卷星导入_Day2.xlsx', day2_data)
    create_wjx_excel('CMport_问卷星导入_Day3.xlsx', day3_data)
    print('\n已按问卷星官方格式重新生成（共3个）。')
